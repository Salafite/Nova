import io
import openpyxl
import pytest
from unittest.mock import MagicMock, patch
from datetime import date
from fastapi.testclient import TestClient

from modules.bi.services.excel_export_service import ExcelExportService, excel_export_service
from modules.bi.services.pdf_export_service import PdfExportService, pdf_export_service
from modules.bi.models.executive_analytics import (
    ExecutiveMarginSummary,
    CategoryMarginResponse,
    CategoryMarginItem,
    SkuMarginResponse,
    SkuMarginItem,
    CustomerProfitabilityResponse,
    CustomerProfitabilityItem,
    QuadrantSummaryItem,
    DeliveryFulfillmentSummaryResponse,
    DeliveryRouteMetricItem,
    WarehouseDeliveryMetricItem,
)
from modules.sales.models.commission import CommissionSummaryItem
from apps.api.main import app


@pytest.fixture
def sample_executive_data():
    summary = ExecutiveMarginSummary(
        period='Monthly',
        gross_sales=120000.0,
        discount_amount=5000.0,
        net_revenue=115000.0,
        cogs=70000.0,
        freight_cost=6000.0,
        gross_profit=39000.0,
        gross_margin_pct=33.91,
        total_orders=180,
        total_customers=50,
        average_order_value=638.89,
        low_margin_order_count=12,
        target_margin_pct=20.0,
        gross_profit_growth_pct=8.5,
    )
    categories = CategoryMarginResponse(
        period='Monthly',
        total_categories=3,
        low_margin_category_count=0,
        items=[
            CategoryMarginItem(
                category_id=1,
                category_name='Beverages & Coffee',
                gross_sales=60000.0,
                discount_amount=2000.0,
                net_revenue=58000.0,
                cogs=32000.0,
                freight_cost=3000.0,
                gross_profit=23000.0,
                gross_margin_pct=39.66,
                revenue_share_pct=50.43,
                units_sold=3500.0,
                order_count=90,
                is_low_margin=False,
                status='Healthy',
            ),
            CategoryMarginItem(
                category_id=2,
                category_name='Bakery & Pastries',
                gross_sales=40000.0,
                discount_amount=2000.0,
                net_revenue=38000.0,
                cogs=25000.0,
                freight_cost=2000.0,
                gross_profit=11000.0,
                gross_margin_pct=28.95,
                revenue_share_pct=33.04,
                units_sold=2000.0,
                order_count=60,
                is_low_margin=False,
                status='Healthy',
            ),
        ],
    )
    skus = SkuMarginResponse(
        period='Monthly',
        total_skus=2,
        low_margin_sku_count=1,
        items=[
            SkuMarginItem(
                product_id=101,
                sku_code='COF-001',
                product_name='Espresso Roast Whole Bean 1kg',
                category_name='Beverages & Coffee',
                brand_name='NovaRoast',
                units_sold=800.0,
                avg_selling_price=45.0,
                unit_cost=22.0,
                gross_sales=36000.0,
                discount_amount=1000.0,
                net_revenue=35000.0,
                cogs=17600.0,
                freight_cost=1800.0,
                gross_profit=15600.0,
                gross_margin_pct=44.57,
                is_low_margin=False,
            ),
            SkuMarginItem(
                product_id=102,
                sku_code='BAK-001',
                product_name='Butter Croissant Bulk (Frozen)',
                category_name='Bakery & Pastries',
                brand_name='NovaBake',
                units_sold=1200.0,
                avg_selling_price=20.0,
                unit_cost=17.5,
                gross_sales=24000.0,
                discount_amount=1000.0,
                net_revenue=23000.0,
                cogs=21000.0,
                freight_cost=1200.0,
                gross_profit=800.0,
                gross_margin_pct=3.48,
                is_low_margin=True,
            ),
        ],
    )
    matrix = CustomerProfitabilityResponse(
        period='Monthly',
        total_customers=3,
        revenue_median_threshold=25000.0,
        margin_threshold_pct=15.0,
        quadrants=[
            QuadrantSummaryItem(
                quadrant='Core Stars',
                quadrant_code='Q1',
                description='High Volume, High Margin Accounts',
                customer_count=1,
                total_net_revenue=60000.0,
                total_gross_profit=25000.0,
                avg_margin_pct=41.67,
                revenue_share_pct=52.17,
                profit_share_pct=64.10,
            ),
            QuadrantSummaryItem(
                quadrant='Volume Risks',
                quadrant_code='Q2',
                description='High Volume, Low Margin Accounts',
                customer_count=1,
                total_net_revenue=40000.0,
                total_gross_profit=3500.0,
                avg_margin_pct=8.75,
                revenue_share_pct=34.78,
                profit_share_pct=8.97,
            ),
        ],
        customers=[
            CustomerProfitabilityItem(
                customer_id=1,
                customer_code='CUST-0001',
                customer_name='Grand Hyatt Hotel',
                sales_rep_name='Sarah Sales',
                order_count=25,
                gross_sales=62000.0,
                discount_amount=2000.0,
                net_revenue=60000.0,
                cogs=32000.0,
                freight_cost=3000.0,
                gross_profit=25000.0,
                gross_margin_pct=41.67,
                average_order_value=2400.0,
                quadrant='Core Stars',
                quadrant_code='Q1',
                recommendation='Protect VIP account and upsell premium lines',
            ),
            CustomerProfitabilityItem(
                customer_id=2,
                customer_code='CUST-0002',
                customer_name='Metro Discount Grocery',
                sales_rep_name='John Rep',
                order_count=35,
                gross_sales=42000.0,
                discount_amount=2000.0,
                net_revenue=40000.0,
                cogs=34500.0,
                freight_cost=2000.0,
                gross_profit=3500.0,
                gross_margin_pct=8.75,
                average_order_value=1142.86,
                quadrant='Volume Risks',
                quadrant_code='Q2',
                recommendation='Renegotiate discounts and enforce minimum order size',
            ),
        ],
    )
    commissions = [
        CommissionSummaryItem(
            sales_rep_id=10,
            sales_rep_name='Sarah Sales',
            sales_rep_email='sarah@novaerp.com',
            total_invoices=25,
            total_collected=60000.0,
            total_gross_margin=25000.0,
            avg_margin_pct=41.67,
            gross_commission=1250.0,
            discount_penalty=40.0,
            net_commission=1210.0,
            paid_commission=800.0,
            pending_commission=410.0,
        )
    ]
    delivery = DeliveryFulfillmentSummaryResponse(
        period='Monthly',
        total_routes=2,
        total_deliveries=180,
        overall_on_time_rate=95.56,
        overall_completion_rate=98.89,
        total_freight_cost=6000.0,
        avg_freight_cost_per_order=33.33,
        routes=[
            DeliveryRouteMetricItem(
                delivery_route='North Metro Express',
                warehouse_name='Main Distribution Center',
                total_deliveries=100,
                completed_deliveries=99,
                on_time_deliveries=96,
                delayed_deliveries=3,
                on_time_delivery_rate=96.97,
                route_completion_rate=99.0,
                total_freight_cost=3200.0,
                avg_freight_per_delivery=32.0,
                total_qty_ordered=5000.0,
                total_qty_shipped=4980.0,
                fulfillment_variance_pct=-0.4,
            )
        ],
    )
    warehouse_metrics = [
        WarehouseDeliveryMetricItem(
            warehouse_id=1,
            warehouse_name='Main Distribution Center',
            location='Industrial Area 1',
            total_deliveries=180,
            completed_deliveries=178,
            on_time_deliveries=172,
            delayed_deliveries=6,
            on_time_delivery_rate=96.63,
            route_completion_rate=98.89,
            total_freight_cost=6000.0,
            avg_freight_per_delivery=33.33,
            total_qty_shipped=8500.0,
        )
    ]
    return summary, categories, skus, matrix, commissions, delivery, warehouse_metrics


class TestExcelExportService:
    """Validates multi-tab structured Excel workbook export generator."""

    def test_excel_workbook_generation_with_all_sheets(self, sample_executive_data):
        summary, categories, skus, matrix, commissions, delivery, warehouse_metrics = sample_executive_data

        exec_svc = MagicMock()
        exec_svc.get_margin_summary.return_value = summary
        exec_svc.get_category_margins.return_value = categories
        exec_svc.get_sku_margins.return_value = skus
        exec_svc.get_period_margin_trends.return_value = MagicMock(items=[])

        cust_svc = MagicMock()
        cust_svc.get_customer_profitability_matrix.return_value = matrix

        comm_svc = MagicMock()
        comm_svc.get_commission_summaries.return_value = commissions

        deliv_svc = MagicMock()
        deliv_svc.get_delivery_fulfillment_summary.return_value = delivery
        deliv_svc.get_warehouse_efficiency.return_value = warehouse_metrics

        service = ExcelExportService(
            executive_service=exec_svc,
            customer_service=cust_svc,
            delivery_service=deliv_svc,
            commission_service=comm_svc,
        )

        excel_buf = service.generate_workbook()
        assert excel_buf is not None
        assert isinstance(excel_buf, io.BytesIO)

        # Inspect workbook contents
        wb = openpyxl.load_workbook(excel_buf)
        expected_sheets = [
            'Executive Summary',
            'Category & SKU Margins',
            'Customer Profitability Matrix',
            'Sales Rep Commissions',
            'Delivery & Logistics',
        ]
        assert wb.sheetnames == expected_sheets

        # Validate Executive Summary sheet
        ws_exec = wb['Executive Summary']
        assert 'NOVA ERP — EXECUTIVE MARGIN & FINANCIAL PERFORMANCE' in ws_exec['A1'].value
        # Check that numeric values are present
        found_sales = False
        for r in range(1, 20):
            for c in range(1, 5):
                if ws_exec.cell(row=r, column=c).value == 120000.0:
                    found_sales = True
                    break
        assert found_sales

        # Validate Category & SKU Margins sheet
        ws_cat = wb['Category & SKU Margins']
        assert 'Beverages & Coffee' in [ws_cat.cell(row=r, column=1).value for r in range(1, 10)]

        # Validate Customer Matrix sheet
        ws_cust = wb['Customer Profitability Matrix']
        assert 'Core Stars' in [ws_cust.cell(row=r, column=1).value for r in range(1, 15)]

        # Validate Commission sheet
        ws_comm = wb['Sales Rep Commissions']
        assert 'Sarah Sales' in [ws_comm.cell(row=r, column=1).value for r in range(1, 15)]

        # Validate Delivery sheet
        ws_del = wb['Delivery & Logistics']
        assert 'North Metro Express' in [ws_del.cell(row=r, column=1).value for r in range(1, 15)]


class TestPdfExportService:
    """Validates board-ready PDF financial report generator."""

    def test_pdf_document_generation(self, sample_executive_data):
        summary, categories, skus, matrix, commissions, delivery, _ = sample_executive_data

        exec_svc = MagicMock()
        exec_svc.get_margin_summary.return_value = summary
        exec_svc.get_category_margins.return_value = categories
        exec_svc.get_sku_margins.return_value = skus

        cust_svc = MagicMock()
        cust_svc.get_customer_profitability_matrix.return_value = matrix

        comm_svc = MagicMock()
        comm_svc.get_commission_summaries.return_value = commissions

        deliv_svc = MagicMock()
        deliv_svc.get_delivery_fulfillment_summary.return_value = delivery

        service = PdfExportService(
            executive_service=exec_svc,
            customer_service=cust_svc,
            delivery_service=deliv_svc,
            commission_service=comm_svc,
        )

        pdf_buf = service.generate_pdf()
        assert pdf_buf is not None
        pdf_bytes = pdf_buf.getvalue()
        assert len(pdf_bytes) > 2000
        assert pdf_bytes.startswith(b'%PDF')


class TestExportControllerEndpoints:
    """Tests streaming REST export endpoints on /api/bi/executive/export."""

    @pytest.fixture(autouse=True)
    def setup_client(self):
        self.client = TestClient(app)

    def _make_auth_header(self, user_id=1, role='Admin', permissions=None):
        from packages.auth.jwt import create_access_token
        token = create_access_token(user_id)
        user_dict = {
            'id': user_id,
            'username': f'user_{role.lower().replace(" ", "_")}_{user_id}',
            'full_name': f'Test {role}',
            'email': f'{role.lower().replace(" ", "_")}@example.com',
            'role': role,
            'permissions': permissions,
            'status': 'Active',
            'business_id': 1,
        }
        return {'Authorization': f'Bearer {token}'}, user_dict

    def test_export_pdf_endpoint(self):
        headers, user = self._make_auth_header(1, 'Admin', permissions=['BI_VIEW'])
        with patch('packages.auth.deps.get_user_by_id', return_value=user), \
             patch('modules.bi.services.pdf_export_service.pdf_export_service.generate_pdf') as mock_pdf:
            fake_buf = io.BytesIO(b'%PDF-1.4 Fake PDF Content')
            mock_pdf.return_value = fake_buf

            resp = self.client.get('/api/bi/executive/export/pdf', headers=headers)
            assert resp.status_code == 200
            assert resp.headers['content-type'] == 'application/pdf'
            assert 'Nova_Executive_Margin_Report_' in resp.headers.get('content-disposition', '')
            assert resp.content.startswith(b'%PDF')

    def test_export_excel_endpoint(self):
        headers, user = self._make_auth_header(1, 'Admin', permissions=['BI_VIEW'])
        with patch('packages.auth.deps.get_user_by_id', return_value=user), \
             patch('modules.bi.services.excel_export_service.excel_export_service.generate_workbook') as mock_excel:
            fake_buf = io.BytesIO(b'PK\x03\x04Fake Excel Content')
            mock_excel.return_value = fake_buf

            resp = self.client.get('/api/bi/executive/export/excel', headers=headers)
            assert resp.status_code == 200
            assert 'spreadsheetml' in resp.headers['content-type']
            assert 'Nova_Executive_Financial_Model_' in resp.headers.get('content-disposition', '')

    def test_export_csv_endpoint(self):
        headers, user = self._make_auth_header(1, 'Admin', permissions=['BI_VIEW'])
        with patch('packages.auth.deps.get_user_by_id', return_value=user), \
             patch('modules.bi.services.executive_analytics_service.executive_analytics_service.get_category_margins') as mock_cat, \
             patch('modules.bi.services.executive_analytics_service.executive_analytics_service.get_margin_summary') as mock_sum:
            mock_cat.return_value = CategoryMarginResponse(
                items=[CategoryMarginItem(category_name='Beverages', gross_sales=5000.0, net_revenue=4800.0, gross_profit=1800.0, gross_margin_pct=37.5)]
            )
            mock_sum.return_value = ExecutiveMarginSummary(
                gross_sales=5000.0, net_revenue=4800.0, gross_profit=1800.0, gross_margin_pct=37.5
            )

            resp = self.client.get('/api/bi/executive/export/csv', headers=headers)
            assert resp.status_code == 200
            assert 'text/csv' in resp.headers['content-type']
            assert b'NOVA ERP' in resp.content
            assert b'Beverages' in resp.content

    def test_export_unauthorized_without_permission(self):
        headers, user = self._make_auth_header(2, 'Viewer', permissions=['PRODUCTS_VIEW'])
        with patch('packages.auth.deps.get_user_by_id', return_value=user):
            resp = self.client.get('/api/bi/executive/export/pdf', headers=headers)
            assert resp.status_code == 403
