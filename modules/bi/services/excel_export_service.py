import io
import logging
from typing import Optional, Dict, Any, Union
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.executive_analytics import (
    ExecutiveAnalyticsFilter,
    ExecutiveMarginSummary,
    CategoryMarginResponse,
    SkuMarginResponse,
    CustomerProfitabilityResponse,
    DeliveryFulfillmentSummaryResponse,
)
from .executive_analytics_service import (
    ExecutiveAnalyticsService,
    executive_analytics_service as default_executive_service,
    resolve_date_range,
)
from .customer_profitability_service import (
    CustomerProfitabilityService,
    customer_profitability_service as default_customer_service,
)
from .delivery_analytics_service import (
    DeliveryAnalyticsService,
    delivery_analytics_service as default_delivery_service,
)
from modules.sales.services.commission_service import (
    CommissionService,
    commission_service as default_commission_service,
)

logger = logging.getLogger(__name__)

# Styling Constants
FONT_FAMILY = "Segoe UI"
HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")  # Royal Blue
SUBHEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
SECTION_TITLE_FONT = Font(name=FONT_FAMILY, size=12, bold=True, color="1E3A8A")
TITLE_FONT = Font(name=FONT_FAMILY, size=16, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name=FONT_FAMILY, size=10, italic=True, color="64748B")
DATA_FONT = Font(name=FONT_FAMILY, size=10, color="1E293B")
BOLD_DATA_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="0F172A")
ALERT_LOW_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="DC2626")  # Red for < 15% margin
ALERT_HIGH_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="16A34A")  # Green for >= 20%
ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
TOTAL_ROW_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
DOUBLE_BOTTOM_SIDE = Side(border_style="double", color="475569")
DATA_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE
)
TOTAL_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=DOUBLE_BOTTOM_SIDE
)

CURRENCY_FORMAT = "$#,##0.00"
PERCENT_FORMAT = "0.0%"
INTEGER_FORMAT = "#,##0"
DECIMAL_FORMAT = "#,##0.00"


def _auto_fit_columns(ws, min_width=12, max_width=45):
    """Adjusts column widths based on maximum contents."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = cell.value
            if val is not None:
                val_str = str(val).split("\n")[0]
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


class ExcelExportService:
    """
    Service for generating comprehensive, board-ready, multi-tab structured Excel workbooks
    for Executive Margin Analytics, Customer Profitability, Commissions, and Logistics.
    """

    def __init__(
        self,
        executive_service: Optional[ExecutiveAnalyticsService] = None,
        customer_service: Optional[CustomerProfitabilityService] = None,
        delivery_service: Optional[DeliveryAnalyticsService] = None,
        commission_service: Optional[CommissionService] = None,
    ):
        self.executive_service = executive_service or default_executive_service
        self.customer_service = customer_service or default_customer_service
        self.delivery_service = delivery_service or default_delivery_service
        self.commission_service = commission_service or default_commission_service

    def _normalize_filter(
        self, filter_input: Union[ExecutiveAnalyticsFilter, Dict[str, Any], None]
    ) -> tuple[ExecutiveAnalyticsFilter, date, date]:
        if filter_input is None:
            flt = ExecutiveAnalyticsFilter()
        elif isinstance(filter_input, dict):
            flt = ExecutiveAnalyticsFilter(**filter_input)
        else:
            flt = filter_input

        start_date, end_date = resolve_date_range(
            period=flt.period,
            date_from=flt.date_from,
            date_to=flt.date_to,
        )
        return flt, start_date, end_date

    def generate_workbook(
        self,
        filters: Union[ExecutiveAnalyticsFilter, Dict[str, Any], None] = None,
        confidentiality_notice: str = "CONFIDENTIAL - BOARD & EXECUTIVE REVIEW ONLY",
        conn=None,
    ) -> io.BytesIO:
        """
        Builds and returns an in-memory multi-tab Excel workbook containing:
        1. Executive Summary & Key Financial KPIs
        2. Category & SKU Margin Breakdown
        3. Customer Profitability Matrix (4-Quadrant)
        4. Sales Rep Commission Ledger
        5. Delivery Route Fulfillment Analytics
        """
        flt, start_date, end_date = self._normalize_filter(filters)

        # Retrieve Data Sources
        summary_kpi = self.executive_service.get_margin_summary(filters=flt, conn=conn)
        category_margins = self.executive_service.get_category_margins(filters=flt, conn=conn)
        sku_margins = self.executive_service.get_sku_margins(filters=flt, limit=150, offset=0, conn=conn)
        margin_trends = self.executive_service.get_period_margin_trends(
            period_type="Monthly", periods_count=12, filters=flt, conn=conn
        )
        customer_matrix = self.customer_service.get_customer_profitability_matrix(filters=flt, conn=conn)
        commissions = self.commission_service.get_commission_summaries(
            period_start=start_date,
            period_end=end_date,
            sales_rep_id=flt.sales_rep_id,
            conn=conn,
        )
        delivery_summary = self.delivery_service.get_delivery_fulfillment_summary(filters=flt, conn=conn)
        warehouse_metrics = self.delivery_service.get_warehouse_efficiency(filters=flt, conn=conn)

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Build Sheets
        self._build_executive_summary_sheet(wb, summary_kpi, margin_trends, flt, start_date, end_date, confidentiality_notice)
        self._build_category_sku_sheet(wb, category_margins, sku_margins)
        self._build_customer_matrix_sheet(wb, customer_matrix)
        self._build_commission_sheet(wb, commissions, start_date, end_date)
        self._build_delivery_analytics_sheet(wb, delivery_summary, warehouse_metrics)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # -----------------------------------------------------------------------
    # Sheet 1: Executive Summary
    # -----------------------------------------------------------------------
    def _build_executive_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        summary: ExecutiveMarginSummary,
        trends,
        flt: ExecutiveAnalyticsFilter,
        start_date: date,
        end_date: date,
        confidentiality_notice: str,
    ):
        ws = wb.create_sheet(title="Executive Summary")
        ws.views.sheetView[0].showGridLines = True

        # Header Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "NOVA ERP — EXECUTIVE MARGIN & FINANCIAL PERFORMANCE"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Period: {summary.period} | Range: {start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = SUBTITLE_FONT

        ws.merge_cells("A3:G3")
        ws["A3"] = confidentiality_notice
        ws["A3"].font = Font(name=FONT_FAMILY, size=9, bold=True, color="991B1B")

        # Section 1: Executive KPI Table
        ws["A5"] = "1. Key Financial Margin Indicators"
        ws["A5"].font = SECTION_TITLE_FONT

        kpi_headers = ["Financial KPI", "Realized Value", "Benchmark / Prev", "Variance / Note"]
        for col_idx, text in enumerate(kpi_headers, start=1):
            cell = ws.cell(row=6, column=col_idx, value=text)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")
            cell.border = DATA_BORDER

        kpi_rows = [
            ("Gross Sales Bookings", summary.gross_sales, CURRENCY_FORMAT, "Top-line invoiced sales volume"),
            ("Customer Discounts Allowed", summary.discount_amount, CURRENCY_FORMAT, "Price concessions & promotional reductions"),
            ("Net Realized Revenue", summary.net_revenue, CURRENCY_FORMAT, "Gross Sales minus Discounts"),
            ("Cost of Goods Sold (COGS)", summary.cogs, CURRENCY_FORMAT, "Product acquisition / production unit costs"),
            ("Freight & Logistics Costs", summary.freight_cost, CURRENCY_FORMAT, "Outbound freight & carrier expenses"),
            ("Realized Gross Profit ($)", summary.gross_profit, CURRENCY_FORMAT, f"Growth: {summary.gross_profit_growth_pct or 0.0:+.1f}% vs prior"),
            ("Gross Profit Margin (%)", summary.gross_margin_pct / 100.0, PERCENT_FORMAT, f"Target: {summary.target_margin_pct:.1f}%"),
            ("Total Completed Orders", summary.total_orders, INTEGER_FORMAT, "Active fulfillment orders"),
            ("Active B2B Customers", summary.total_customers, INTEGER_FORMAT, "Customers with order transactions"),
            ("Average Order Value (AOV)", summary.average_order_value, CURRENCY_FORMAT, "Net Revenue per order"),
            ("Low-Margin Order Count (<15%)", summary.low_margin_order_count, INTEGER_FORMAT, "Orders requiring pricing review"),
        ]

        for row_idx, (kpi_name, val, num_fmt, note) in enumerate(kpi_rows, start=7):
            c1 = ws.cell(row=row_idx, column=1, value=kpi_name)
            c2 = ws.cell(row=row_idx, column=2, value=val)
            c3 = ws.cell(row=row_idx, column=3, value="")
            c4 = ws.cell(row=row_idx, column=4, value=note)

            c1.font = BOLD_DATA_FONT if "Gross Profit" in kpi_name else DATA_FONT
            c2.font = BOLD_DATA_FONT if "Gross Profit" in kpi_name else DATA_FONT
            c2.number_format = num_fmt
            c2.alignment = Alignment(horizontal="right")
            c4.font = SUBTITLE_FONT

            # Highlight low gross margin
            if kpi_name == "Gross Profit Margin (%)":
                if summary.gross_margin_pct < 15.0:
                    c2.font = ALERT_LOW_FONT
                else:
                    c2.font = ALERT_HIGH_FONT

            for c in (c1, c2, c3, c4):
                c.border = DATA_BORDER
                if row_idx % 2 == 0:
                    c.fill = ALT_ROW_FILL

        # Section 2: Historical Trends Table
        trend_start_row = len(kpi_rows) + 9
        ws.cell(row=trend_start_row, column=1, value="2. Monthly Gross Margin Historical Performance").font = SECTION_TITLE_FONT

        trend_headers = ["Period", "Gross Sales", "Discounts", "Net Revenue", "COGS", "Freight Cost", "Gross Profit", "Margin %", "Orders"]
        for col_idx, text in enumerate(trend_headers, start=1):
            cell = ws.cell(row=trend_start_row + 1, column=col_idx, value=text)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")
            cell.border = DATA_BORDER

        trend_items = getattr(trends, "items", []) if trends else []
        for idx, t in enumerate(trend_items, start=trend_start_row + 2):
            ws.cell(row=idx, column=1, value=t.period_label).alignment = Alignment(horizontal="left")
            c_sales = ws.cell(row=idx, column=2, value=t.gross_sales)
            c_disc = ws.cell(row=idx, column=3, value=t.discount_amount)
            c_rev = ws.cell(row=idx, column=4, value=t.net_revenue)
            c_cogs = ws.cell(row=idx, column=5, value=t.cogs)
            c_frt = ws.cell(row=idx, column=6, value=t.freight_cost)
            c_profit = ws.cell(row=idx, column=7, value=t.gross_profit)
            c_margin = ws.cell(row=idx, column=8, value=t.gross_margin_pct / 100.0)
            c_ord = ws.cell(row=idx, column=9, value=t.order_count)

            for c in (c_sales, c_disc, c_rev, c_cogs, c_frt, c_profit):
                c.number_format = CURRENCY_FORMAT
                c.alignment = Alignment(horizontal="right")
            c_margin.number_format = PERCENT_FORMAT
            c_margin.alignment = Alignment(horizontal="right")
            if t.gross_margin_pct < 15.0:
                c_margin.font = ALERT_LOW_FONT
            c_ord.number_format = INTEGER_FORMAT
            c_ord.alignment = Alignment(horizontal="right")

            for col in range(1, 10):
                cell = ws.cell(row=idx, column=col)
                cell.font = DATA_FONT if cell.font == Font() else cell.font
                cell.border = DATA_BORDER
                if idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        _auto_fit_columns(ws)

    # -----------------------------------------------------------------------
    # Sheet 2: Category & SKU Margins
    # -----------------------------------------------------------------------
    def _build_category_sku_sheet(
        self,
        wb: openpyxl.Workbook,
        category_res: CategoryMarginResponse,
        sku_res: SkuMarginResponse,
    ):
        ws = wb.create_sheet(title="Category & SKU Margins")
        ws.views.sheetView[0].showGridLines = True

        ws["A1"] = "Product Category Margin Optimization"
        ws["A1"].font = TITLE_FONT

        # Category Table
        cat_headers = [
            "Category", "Gross Sales", "Discounts", "Net Revenue", "COGS",
            "Freight Cost", "Gross Profit", "Margin %", "Rev Share %", "Units Sold", "Orders", "Status"
        ]
        for col_idx, text in enumerate(cat_headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 12) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row = 4
        cat_items = category_res.items or []
        for cat in cat_items:
            ws.cell(row=curr_row, column=1, value=cat.category_name).alignment = Alignment(horizontal="left")
            c_sales = ws.cell(row=curr_row, column=2, value=cat.gross_sales)
            c_disc = ws.cell(row=curr_row, column=3, value=cat.discount_amount)
            c_rev = ws.cell(row=curr_row, column=4, value=cat.net_revenue)
            c_cogs = ws.cell(row=curr_row, column=5, value=cat.cogs)
            c_frt = ws.cell(row=curr_row, column=6, value=cat.freight_cost)
            c_profit = ws.cell(row=curr_row, column=7, value=cat.gross_profit)
            c_margin = ws.cell(row=curr_row, column=8, value=cat.gross_margin_pct / 100.0)
            c_share = ws.cell(row=curr_row, column=9, value=cat.revenue_share_pct / 100.0)
            c_units = ws.cell(row=curr_row, column=10, value=cat.units_sold)
            c_ord = ws.cell(row=curr_row, column=11, value=cat.order_count)
            c_status = ws.cell(row=curr_row, column=12, value=cat.status)

            for c in (c_sales, c_disc, c_rev, c_cogs, c_frt, c_profit):
                c.number_format = CURRENCY_FORMAT
                c.alignment = Alignment(horizontal="right")
            c_margin.number_format = PERCENT_FORMAT
            c_margin.alignment = Alignment(horizontal="right")
            c_share.number_format = PERCENT_FORMAT
            c_share.alignment = Alignment(horizontal="right")
            c_units.number_format = DECIMAL_FORMAT
            c_units.alignment = Alignment(horizontal="right")
            c_ord.number_format = INTEGER_FORMAT
            c_ord.alignment = Alignment(horizontal="right")
            c_status.alignment = Alignment(horizontal="center")

            if cat.is_low_margin:
                c_margin.font = ALERT_LOW_FONT
                c_status.font = ALERT_LOW_FONT

            for col in range(1, 13):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        # SKU Section
        curr_row += 2
        ws.cell(row=curr_row, column=1, value="SKU-Level Profitability & Cost Breakdown").font = SECTION_TITLE_FONT
        curr_row += 1

        sku_headers = [
            "SKU Code", "Product Name", "Category", "Brand", "Units Sold", "Avg Price",
            "Unit Cost", "Gross Sales", "Net Revenue", "COGS", "Freight Cost", "Gross Profit", "Margin %", "Low Margin Alert"
        ]
        for col_idx, text in enumerate(sku_headers, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=text)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2, 3, 4, 14) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row += 1
        sku_items = sku_res.items or []
        for sku in sku_items:
            ws.cell(row=curr_row, column=1, value=sku.sku_code or f"SKU-{sku.product_id}").alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=2, value=sku.product_name).alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=3, value=sku.category_name or "").alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=4, value=sku.brand_name or "").alignment = Alignment(horizontal="left")
            c_units = ws.cell(row=curr_row, column=5, value=sku.units_sold)
            c_price = ws.cell(row=curr_row, column=6, value=sku.avg_selling_price)
            c_cost = ws.cell(row=curr_row, column=7, value=sku.unit_cost)
            c_sales = ws.cell(row=curr_row, column=8, value=sku.gross_sales)
            c_rev = ws.cell(row=curr_row, column=9, value=sku.net_revenue)
            c_cogs = ws.cell(row=curr_row, column=10, value=sku.cogs)
            c_frt = ws.cell(row=curr_row, column=11, value=sku.freight_cost)
            c_profit = ws.cell(row=curr_row, column=12, value=sku.gross_profit)
            c_margin = ws.cell(row=curr_row, column=13, value=sku.gross_margin_pct / 100.0)
            c_alert = ws.cell(row=curr_row, column=14, value="LOW MARGIN (<15%)" if sku.is_low_margin else "OK")

            c_units.number_format = DECIMAL_FORMAT
            c_units.alignment = Alignment(horizontal="right")
            for c in (c_price, c_cost, c_sales, c_rev, c_cogs, c_frt, c_profit):
                c.number_format = CURRENCY_FORMAT
                c.alignment = Alignment(horizontal="right")
            c_margin.number_format = PERCENT_FORMAT
            c_margin.alignment = Alignment(horizontal="right")
            c_alert.alignment = Alignment(horizontal="center")

            if sku.is_low_margin:
                c_margin.font = ALERT_LOW_FONT
                c_alert.font = ALERT_LOW_FONT

            for col in range(1, 15):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        _auto_fit_columns(ws)

    # -----------------------------------------------------------------------
    # Sheet 3: Customer Profitability Matrix
    # -----------------------------------------------------------------------
    def _build_customer_matrix_sheet(
        self,
        wb: openpyxl.Workbook,
        matrix_res: CustomerProfitabilityResponse,
    ):
        ws = wb.create_sheet(title="Customer Profitability Matrix")
        ws.views.sheetView[0].showGridLines = True

        ws["A1"] = "Customer Profitability Matrix (4-Quadrant Strategic Segmentation)"
        ws["A1"].font = TITLE_FONT

        # Section 1: Quadrants Summary
        quad_headers = [
            "Quadrant", "Code", "Strategic Description", "Accounts",
            "Net Revenue", "Gross Profit", "Avg Margin %", "Rev Share %", "Profit Share %"
        ]
        for col_idx, text in enumerate(quad_headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2, 3) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row = 4
        for q in matrix_res.quadrants or []:
            ws.cell(row=curr_row, column=1, value=q.quadrant).alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=2, value=q.quadrant_code).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=3, value=q.description).alignment = Alignment(horizontal="left")
            c_cnt = ws.cell(row=curr_row, column=4, value=q.customer_count)
            c_rev = ws.cell(row=curr_row, column=5, value=q.total_net_revenue)
            c_profit = ws.cell(row=curr_row, column=6, value=q.total_gross_profit)
            c_margin = ws.cell(row=curr_row, column=7, value=q.avg_margin_pct / 100.0)
            c_rshare = ws.cell(row=curr_row, column=8, value=q.revenue_share_pct / 100.0)
            c_pshare = ws.cell(row=curr_row, column=9, value=q.profit_share_pct / 100.0)

            c_cnt.number_format = INTEGER_FORMAT
            c_cnt.alignment = Alignment(horizontal="right")
            c_rev.number_format = CURRENCY_FORMAT
            c_rev.alignment = Alignment(horizontal="right")
            c_profit.number_format = CURRENCY_FORMAT
            c_profit.alignment = Alignment(horizontal="right")
            c_margin.number_format = PERCENT_FORMAT
            c_margin.alignment = Alignment(horizontal="right")
            c_rshare.number_format = PERCENT_FORMAT
            c_rshare.alignment = Alignment(horizontal="right")
            c_pshare.number_format = PERCENT_FORMAT
            c_pshare.alignment = Alignment(horizontal="right")

            for col in range(1, 10):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        # Section 2: Customer Account Rankings
        curr_row += 2
        ws.cell(row=curr_row, column=1, value="Ranked Customer Accounts Ledger").font = SECTION_TITLE_FONT
        curr_row += 1

        cust_headers = [
            "Customer Code", "Customer Name", "Sales Rep", "Quadrant", "Orders",
            "Gross Sales", "Discounts", "Net Revenue", "COGS", "Freight Cost",
            "Gross Profit", "Margin %", "AOV", "Strategic Recommendation"
        ]
        for col_idx, text in enumerate(cust_headers, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=text)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2, 3, 4, 14) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row += 1
        for cust in matrix_res.customers or []:
            ws.cell(row=curr_row, column=1, value=cust.customer_code or f"CUST-{cust.customer_id:04d}").alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=2, value=cust.customer_name).alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=3, value=cust.sales_rep_name or "Unassigned").alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=4, value=f"{cust.quadrant} ({cust.quadrant_code})").alignment = Alignment(horizontal="left")
            c_ord = ws.cell(row=curr_row, column=5, value=cust.order_count)
            c_sales = ws.cell(row=curr_row, column=6, value=cust.gross_sales)
            c_disc = ws.cell(row=curr_row, column=7, value=cust.discount_amount)
            c_rev = ws.cell(row=curr_row, column=8, value=cust.net_revenue)
            c_cogs = ws.cell(row=curr_row, column=9, value=cust.cogs)
            c_frt = ws.cell(row=curr_row, column=10, value=cust.freight_cost)
            c_profit = ws.cell(row=curr_row, column=11, value=cust.gross_profit)
            c_margin = ws.cell(row=curr_row, column=12, value=cust.gross_margin_pct / 100.0)
            c_aov = ws.cell(row=curr_row, column=13, value=cust.average_order_value)
            ws.cell(row=curr_row, column=14, value=cust.recommendation or "").alignment = Alignment(horizontal="left")

            c_ord.number_format = INTEGER_FORMAT
            c_ord.alignment = Alignment(horizontal="right")
            for c in (c_sales, c_disc, c_rev, c_cogs, c_frt, c_profit, c_aov):
                c.number_format = CURRENCY_FORMAT
                c.alignment = Alignment(horizontal="right")
            c_margin.number_format = PERCENT_FORMAT
            c_margin.alignment = Alignment(horizontal="right")

            if cust.quadrant_code == "Q4" or cust.gross_margin_pct < 10.0:
                c_margin.font = ALERT_LOW_FONT
            elif cust.quadrant_code == "Q1":
                c_margin.font = ALERT_HIGH_FONT

            for col in range(1, 15):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        _auto_fit_columns(ws)

    # -----------------------------------------------------------------------
    # Sheet 4: Sales Rep Commissions
    # -----------------------------------------------------------------------
    def _build_commission_sheet(
        self,
        wb: openpyxl.Workbook,
        commissions: list,
        start_date: date,
        end_date: date,
    ):
        ws = wb.create_sheet(title="Sales Rep Commissions")
        ws.views.sheetView[0].showGridLines = True

        ws["A1"] = "Sales Representative Collected Margin Commission Ledger"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Calculated on Collected Cash & Realized Gross Profit | {start_date} to {end_date}"
        ws["A2"].font = SUBTITLE_FONT

        comm_headers = [
            "Sales Rep Name", "Email", "Invoices", "Total Collected Cash", "Realized Gross Margin",
            "Realized Margin %", "Gross Commission", "Discount Penalty", "Net Commission Payable",
            "Paid Commission", "Pending Balance"
        ]
        for col_idx, text in enumerate(comm_headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row = 5
        tot_collected = 0.0
        tot_margin = 0.0
        tot_gross_comm = 0.0
        tot_penalty = 0.0
        tot_net_comm = 0.0
        tot_paid = 0.0
        tot_pending = 0.0

        for c in commissions:
            ws.cell(row=curr_row, column=1, value=c.sales_rep_name).alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=2, value=c.sales_rep_email or "").alignment = Alignment(horizontal="left")
            c_inv = ws.cell(row=curr_row, column=3, value=c.total_invoices)
            c_col = ws.cell(row=curr_row, column=4, value=c.total_collected)
            c_mar = ws.cell(row=curr_row, column=5, value=c.total_gross_margin)
            c_mar_pct = ws.cell(row=curr_row, column=6, value=c.avg_margin_pct / 100.0)
            c_gcomm = ws.cell(row=curr_row, column=7, value=c.gross_commission)
            c_pen = ws.cell(row=curr_row, column=8, value=c.discount_penalty)
            c_net = ws.cell(row=curr_row, column=9, value=c.net_commission)
            c_paid = ws.cell(row=curr_row, column=10, value=c.paid_commission)
            c_pend = ws.cell(row=curr_row, column=11, value=c.pending_commission)

            c_inv.number_format = INTEGER_FORMAT
            c_inv.alignment = Alignment(horizontal="right")
            for cell in (c_col, c_mar, c_gcomm, c_pen, c_net, c_paid, c_pend):
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = Alignment(horizontal="right")
            c_mar_pct.number_format = PERCENT_FORMAT
            c_mar_pct.alignment = Alignment(horizontal="right")

            tot_collected += c.total_collected
            tot_margin += c.total_gross_margin
            tot_gross_comm += c.gross_commission
            tot_penalty += c.discount_penalty
            tot_net_comm += c.net_commission
            tot_paid += c.paid_commission
            tot_pending += c.pending_commission

            for col in range(1, 12):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        # Summary Row
        t_label = ws.cell(row=curr_row, column=1, value="TOTAL ALL REPS")
        t_label.font = BOLD_DATA_FONT
        t_label.alignment = Alignment(horizontal="left")
        ws.cell(row=curr_row, column=2, value="")
        ws.cell(row=curr_row, column=3, value="")

        t_col = ws.cell(row=curr_row, column=4, value=tot_collected)
        t_mar = ws.cell(row=curr_row, column=5, value=tot_margin)
        t_mar_pct = ws.cell(row=curr_row, column=6, value=(tot_margin / tot_collected) if tot_collected > 0 else 0.0)
        t_gcomm = ws.cell(row=curr_row, column=7, value=tot_gross_comm)
        t_pen = ws.cell(row=curr_row, column=8, value=tot_penalty)
        t_net = ws.cell(row=curr_row, column=9, value=tot_net_comm)
        t_paid = ws.cell(row=curr_row, column=10, value=tot_paid)
        t_pend = ws.cell(row=curr_row, column=11, value=tot_pending)

        for cell in (t_col, t_mar, t_gcomm, t_pen, t_net, t_paid, t_pend):
            cell.number_format = CURRENCY_FORMAT
            cell.font = BOLD_DATA_FONT
            cell.alignment = Alignment(horizontal="right")
        t_mar_pct.number_format = PERCENT_FORMAT
        t_mar_pct.font = BOLD_DATA_FONT
        t_mar_pct.alignment = Alignment(horizontal="right")

        for col in range(1, 12):
            c = ws.cell(row=curr_row, column=col)
            c.border = TOTAL_BORDER
            c.fill = TOTAL_ROW_FILL

        _auto_fit_columns(ws)

    # -----------------------------------------------------------------------
    # Sheet 5: Delivery & Fulfillment Analytics
    # -----------------------------------------------------------------------
    def _build_delivery_analytics_sheet(
        self,
        wb: openpyxl.Workbook,
        delivery_res: DeliveryFulfillmentSummaryResponse,
        warehouse_metrics: list,
    ):
        ws = wb.create_sheet(title="Delivery & Logistics")
        ws.views.sheetView[0].showGridLines = True

        ws["A1"] = "Delivery Route Fulfillment & Freight Efficiency"
        ws["A1"].font = TITLE_FONT

        # Route table
        route_headers = [
            "Delivery Route", "Warehouse", "Total Deliveries", "Completed",
            "On-Time", "Delayed", "On-Time Rate %", "Completion Rate %",
            "Total Freight Cost", "Avg Freight / Delivery", "Qty Ordered", "Qty Shipped", "Variance %"
        ]
        for col_idx, text in enumerate(route_headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row = 4
        for r in delivery_res.routes or []:
            ws.cell(row=curr_row, column=1, value=r.delivery_route).alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=2, value=r.warehouse_name or "").alignment = Alignment(horizontal="left")
            c_tot = ws.cell(row=curr_row, column=3, value=r.total_deliveries)
            c_comp = ws.cell(row=curr_row, column=4, value=r.completed_deliveries)
            c_ont = ws.cell(row=curr_row, column=5, value=r.on_time_deliveries)
            c_del = ws.cell(row=curr_row, column=6, value=r.delayed_deliveries)
            c_otd = ws.cell(row=curr_row, column=7, value=r.on_time_delivery_rate / 100.0)
            c_cmp = ws.cell(row=curr_row, column=8, value=r.route_completion_rate / 100.0)
            c_frt = ws.cell(row=curr_row, column=9, value=r.total_freight_cost)
            c_avgf = ws.cell(row=curr_row, column=10, value=r.avg_freight_per_delivery)
            c_ord = ws.cell(row=curr_row, column=11, value=r.total_qty_ordered)
            c_shp = ws.cell(row=curr_row, column=12, value=r.total_qty_shipped)
            c_var = ws.cell(row=curr_row, column=13, value=r.fulfillment_variance_pct / 100.0)

            for c in (c_tot, c_comp, c_ont, c_del):
                c.number_format = INTEGER_FORMAT
                c.alignment = Alignment(horizontal="right")
            for c in (c_otd, c_cmp, c_var):
                c.number_format = PERCENT_FORMAT
                c.alignment = Alignment(horizontal="right")
            for c in (c_frt, c_avgf):
                c.number_format = CURRENCY_FORMAT
                c.alignment = Alignment(horizontal="right")
            for c in (c_ord, c_shp):
                c.number_format = DECIMAL_FORMAT
                c.alignment = Alignment(horizontal="right")

            if r.on_time_delivery_rate < 85.0:
                c_otd.font = ALERT_LOW_FONT
            elif r.on_time_delivery_rate >= 95.0:
                c_otd.font = ALERT_HIGH_FONT

            for col in range(1, 14):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        # Section 2: Warehouse Dispatch Efficiency
        curr_row += 2
        ws.cell(row=curr_row, column=1, value="Origin Warehouse Dispatch & Efficiency").font = SECTION_TITLE_FONT
        curr_row += 1

        wh_headers = [
            "Warehouse Name", "Location", "Total Deliveries", "Completed",
            "On-Time", "Delayed", "On-Time Rate %", "Completion Rate %", "Total Freight Cost", "Avg Freight", "Qty Shipped"
        ]
        for col_idx, text in enumerate(wh_headers, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=text)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 2) else "right", vertical="center")
            cell.border = DATA_BORDER

        curr_row += 1
        for w in warehouse_metrics or []:
            ws.cell(row=curr_row, column=1, value=w.warehouse_name).alignment = Alignment(horizontal="left")
            ws.cell(row=curr_row, column=2, value=w.location or "").alignment = Alignment(horizontal="left")
            c_tot = ws.cell(row=curr_row, column=3, value=w.total_deliveries)
            c_comp = ws.cell(row=curr_row, column=4, value=w.completed_deliveries)
            c_ont = ws.cell(row=curr_row, column=5, value=w.on_time_deliveries)
            c_del = ws.cell(row=curr_row, column=6, value=w.delayed_deliveries)
            c_otd = ws.cell(row=curr_row, column=7, value=w.on_time_delivery_rate / 100.0)
            c_cmp = ws.cell(row=curr_row, column=8, value=w.route_completion_rate / 100.0)
            c_frt = ws.cell(row=curr_row, column=9, value=w.total_freight_cost)
            c_avgf = ws.cell(row=curr_row, column=10, value=w.avg_freight_per_delivery)
            c_shp = ws.cell(row=curr_row, column=11, value=w.total_qty_shipped)

            for c in (c_tot, c_comp, c_ont, c_del):
                c.number_format = INTEGER_FORMAT
                c.alignment = Alignment(horizontal="right")
            for c in (c_otd, c_cmp):
                c.number_format = PERCENT_FORMAT
                c.alignment = Alignment(horizontal="right")
            for c in (c_frt, c_avgf):
                c.number_format = CURRENCY_FORMAT
                c.alignment = Alignment(horizontal="right")
            c_shp.number_format = DECIMAL_FORMAT
            c_shp.alignment = Alignment(horizontal="right")

            for col in range(1, 12):
                cell = ws.cell(row=curr_row, column=col)
                cell.border = DATA_BORDER
                if curr_row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

            curr_row += 1

        _auto_fit_columns(ws)


# Default singleton instance
excel_export_service = ExcelExportService()
